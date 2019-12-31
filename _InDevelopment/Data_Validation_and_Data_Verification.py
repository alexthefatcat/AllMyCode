# -*- coding: utf-8 -*-"""Created on Wed Feb 28 10:49:00 2018@author: milroa1
#%%
import pandas as pd
"""   if the input is a certain type then change it   """
for inp in ["one",1,"Two"]:
    #inp = "odne"
    out = None
    
    if type(inp) is str:
        inp=inp.title().strip()
        out = dict(zip("Zero One Two Three Four Five Six Seven Eight Nine Ten".split(),range(11))).get(inp,inp)

    print(out,"\t:",inp)
    
    #numeric
    if type(inp) in [bool, float, int, complex]:
            inp=float(inp)
            print("-----------numberic------------------")        
    
    
    if isinstance(inp, pd.DataFrame) or isinstance(inp,pd.Series):
        print("pandas object")
    

#%%    
###################################################################
"If you want to remove leading and ending spaces, use str.strip():"

sentence = ' hello  apple'
sentence.strip()
#>>> 'hello  apple'
##########################################################
"If you want to remove all spaces, use str.replace():"

sentence = ' hello  apple'
sentence.replace(" ", "")
#>>> 'helloapple'
###########################################################
"If you want to remove duplicated spaces, use str.split():"

sentence = ' hello  apple'
" ".join(sentence.split())
#>>> 'hello apple'
#############################################################
#%%    
"""     Multi-Processing         """
from multiprocessing.dummy import Pool as ThreadPool 

def write(i, x):
    print(i, "---", x)

a = ["1","2","3"]
b = ["4","5","6"] 

pool = ThreadPool(2)
pool.starmap(write, zip(a,b)) 
pool.close() 
pool.join()
#%%    
df.dtypes


#%%            Data_Validator  
"""   

1) check the extension of the filepath
2) read in the data
3) select the data into a smaller table
  > find a string or value in the dataframe and how it relates to the beggining
  > same but with a rough location> "find uk in A6-B8"
  > hardcoder top left botterm right corners
    > end with last value as option
    >>> what to do missing values fill in with one before
4) are the value in the columns the right type if not then change them int> float type of operation
5) check column names and row names? also if there in the right order
   > make sure all have differnt values
6) are there nans;missing values; any values suprissingly large
7) 
"""


#%%    Usefull Functions


def find_location_of_value_or_function_in_df(df, val, mode=["basic", "iloc", "all", "excel"][0], function_value=[True, False, "string_contaions_all"][0] ):

            if not mode           in ["basic", "iloc", "all", "excel"]:
                print("Error mode has to be one of ['basic', 'iloc', 'all', 'excel']")
                
            if not function_value in [True,False,"string_contaions_all"]:
                print("Error mode has to be one of [True, False, 'string_contaions_all']")              
                
            def num2exc(no):return("abcdefghijklmnopqrstuvwxyz"[no])
            
            import collections,numpy as np
            ## check if val is a function or not
            if isinstance(val, collections.Callable):  
                def func(x):
                    try:
                        m = val(x)
                        if function_value_is_True:
                            m = m in [True,1]
                    except TypeError:
                        m = 0#False
                    return(m)
            else:
                if type(val) is list:
                     if function_value=="string_contaions_all" :     
                                                                      def func(x): return(all([ m in x for m in val]))
                     else :                                          
                                                                      def func(x): return(x in val)
                else:                                                
                                                                      def func(x): return(x==val)
                       
            df_        = df.applymap(func)       
            values     = df_.values
            rows, cols = np.where(values)
            
            if mode in ["basic"] :   
                out = [ [df.index[r],  df.columns[c], df.iloc[r,c]]                     for r,c in zip(rows, cols)] 
                
            if mode in [ "all" ] : 
                out = [ [df.index[r],  df.columns[c],r,c,df_.iloc[r,c], df.iloc[r,c]]   for r,c in zip(rows, cols)] 
                
            if mode in ["iloc" ] : 
                out = [ [         r ,             c , df.iloc[r,c]]                     for r,c in zip(rows, cols)] 

            if mode in ["excel"] : 
                out = [ [  num2exc(no)(c) +str(r+1), df.iloc[r,c]]                       for r,c in zip(rows, cols)]                 
                   
            return( out)


def quick_grab_table(df, row_col, no_columns, no_rows):
            """a function which segements out a table from a dataframe created from csv or excel spreadsheet    
            
            :param df: The Dataframe to extract the table from
            :type  df: DataFrame
            :param row_col: The location in the top left corner in the iloc format, [0,0] is most left top location
            :type  row_col: list
            :param no_columns: The number of columns in the dataframes you want extracted out
            :type  no_columns: int
            :param no_rows: The number of rows in the dataframes you want extracted out
            :type  no_rows: int
            :returns: Dataframe from the relevent input values
            """
            df_out         = df.iloc[ [ row_col[0] + 1 + i for i in range(no_rows)], [row_col[1] + 1 + i for i in range(no_columns)]] 
            df_out.columns = df.iloc[   row_col[0]                                 , [row_col[1] + 1 + i for i in range(no_columns)]]
            df_out.index =   df.iloc[ [ row_col[0] + 1 + i for i in range(no_rows)],  row_col[1]                                   ]
            return(df_out)  


def extract_table_by_corners(dataframe,excel_corners,remove_spaces_in_column=True):
       """
       This function extracts a dataframe from another when you input a string with the excel co-ordinates.
       Example:  HERD_df = extract_table_by_corners(HERD_df_temp,"A6-F12")
       
       :param dataframe: The Dataframe to extract the table from
       :type  dataframe: DataFrame
       :param excel_corners: The string containing the corners eg "A6-F12"
       :type  excel_corners: String  
       :returns: Dataframe from the relevent input values
       """
       
       TEMP=excel_corners.split("-")
       
       def letter2num(let): 
           letters_= [(n+m).replace("_","") for n in "_ABCDEFGHIJKLMNOPQRSTUVWXYZ" for m in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]
           return(letters_.index(let))
           
       if len(TEMP)==2:
              print("*")
              top_left_corner,bot_righ_corner=TEMP[0],TEMP[1]
              tl_let = letter2num("".join([n for n in top_left_corner if not n in "0123456789"]))
              br_let = letter2num("".join([n for n in bot_righ_corner if not n in "0123456789"]))
              tl_num =        int("".join([n for n in top_left_corner if     n in "0123456789"]))
              br_num =        int("".join([n for n in bot_righ_corner if     n in "0123456789"]))
    
              df_out         = dataframe.iloc[ (tl_num+1):(br_num), (tl_let+1):(br_let+1) ] 
              df_out.index =   dataframe.iloc[ (tl_num+1):(br_num),  tl_let               ]
              df_out.columns = dataframe.iloc[  tl_num            , (tl_let+1):(br_let+1) ]
              
              if remove_spaces_in_column:
                   df_out.columns =  [column_name.replace(" ","") for column_name in  df_out.columns ]
       else :
            print("should be in the format 'A4-G8' the input does not contain one '-' ")
            df_out=[]
            
       return(df_out)  



def str_match_dont_match(x,match,dontmatch=[]):
            out= False
            if type(x) is str:
                if all(n in x for n in match):
                   if all(not n in x for n in dontmatch):
                        out=True
            return(out)          
                    
loc = find_location_of_value_or_function_in_df(df, lambda x:str_match_dont_match(x,["Ref"],["References"]), iloc=True)[0][0]

# elmi
examp_df = examp_df.loc[:,[col for col in examp_df.columns if not col in [""] ]]  
examp_df = examp_df.astype(float).astype(int)




def CSV_2_DataFrame(filepath,start=12):
    import csv
    with open(filepath, 'r') as f:
        reader = csv.reader(f) 
        out  = [row for row in reader]
        
    mout = max(len(row) for row in out)

    out=out[start:]
    out2=out.copy()
    for i, row in enumerate(out[1:],1):
        if not len(row)==len(out[0]):
            out2[i]=out2[i]+[[]*len(out[0])]
            out2[i]=out2[i][:len(out[0])]
        if row[0]=="":
           out2[i][0]=out2[i-1][0]
               
        df_out=pd.DataFrame(out2[1:],columns=out2[0])
        df_out.iloc[:,[2  ]]=df_out.iloc[:,[ 2 ]].astype(float).astype( int )
        
        return(df_out)



def print_object_info(cell):
  out={}
  from copy import copy
  for n in dir(cell):
    v="\n"
    if n[0] !="_":
       exec("print(n,v,cell."+n+",v+v)")
       exec("out[n]=copy(cell."+n+")")
  return(out)

import pandas as pd
_df=pd.DataFrame
info = print_object_info(_df)

###################################################################################
## circle thing
import math

tau=2*math.pi
r,m=5,5

z=360
#r,m=z/tau,z
r,m=z/tau,z

dist_sum=0
for n in range(m):
  n=n+0.5
  circ=math.sin((n/m)*tau)
  dist=r*r*(1+circ)
  dist2=dist*dist
  dist_sum=dist_sum+(1/(dist))
  print(n,m,circ,dist,dist2,dist_sum)
#########################################################################################

la=lambda x,y:x*y+x

def apply_two_df(df1,df2,func):
    cols=df1.columns
    cols2=[str(type(n))[8]+str(n) for n in cols]
    df1.columns=["A_"+n for n in cols2]
    df2.columns=["B_"+n for n in cols2]
    df =  pd.concat([df1, df2], axis=1)
    df_out=pd.DataFrame(columns=cols2)
    for col in cols2:
        func2=lambda row: func(row["A_"+col], row["B_"+col])
        df_out[col]=df.apply(func2, axis=1)
    df_out.columns=cols
    return(df_out)

#########################################################################################
note for dataframe convert ot float then int to change its values form an object numeric option
#########################################################################################
def type_letter(obj,len_=1):
    str_=str(type(obj))[8:-2]
    print(str_)
    if len_>0:
        str_=str_[:len_]
    return(str_)

for l in [0,1,2,3]:
    for n in [1,1.22,"ddsds"]:
       print(n,type_letter(n,l),l)
#########################################################################################


for attr, value in C.__dict__.iteritems():

#########################################################################################
#Copy sheet from worksheet IN OPENPYXL
import openpyxl
Filename="tables.xlsx"
Sheetname="Sheet_1"

    ws_new    = wb_new.create_sheet("NEW-COPY")  
    wbs_old   = openpyxl.load_workbook( Filename )
    ws_old    = wbs_old[Sheetname]
    
    from copy import copy

    def copy_object_openpyxl(cell_1, cell_2):                        
        for n in dir(cell_1):
           #if n[0] not in ["_"]:
             #if n not in ["internal_value","value","coordinate","style_id","row","parent","column","col_idx"]:
              if n in ["alignment","border","fill","font","number_format","value"]:##,"style" doesnt work dont know why
                 exec(  "cell_1."+n+"= copy(cell_2."+n +")"  )#"cell_1.fill= copy(cell_2.fill)
        return(cell_2)
        
    for rows in ws_old.rows:
        for cell in rows:
            ws_new[cell.coordinate].value         = cell.value
            copy_object_openpyxl(ws_new[cell.coordinate], cell)

    for co in ws_old.merged_cell_ranges:
        ws_new.merge_cells(str(co)) 
        co_merge=str(co).split(":")
        
        for row in ws_new[co_merge[0]:co_merge[1]]:
            for cell_obj in row:
               cell_loc=str(cell_obj).split(".")[-1][:-1] #"<Cell 'Table 6'.E28>" to  E28
               ws_new[cell_loc].border = copy(ws_old[co_merge[0]].border)    



#########################################################################################



























